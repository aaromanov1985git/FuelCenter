"""
Роутер для работы с транзакциями
"""
from fastapi import APIRouter, UploadFile, File, Depends, Query, HTTPException, Request
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.database import get_db
from app.logger import logger
from app.config import get_settings
from app.models import Transaction, Provider, UploadPeriodLock, ProviderTemplate, User
from app.schemas import (
    TransactionResponse, TransactionListResponse, FileUploadResponse
)
from app.services import detect_provider_and_template
from app.services.transaction_service import TransactionService
from app.services.excel_processor import ExcelProcessor
from app.services.transaction_batch_processor import TransactionBatchProcessor
from app.services.api_provider_service import ApiProviderService
from app.services.upload_event_service import UploadEventService
from app.utils import (
    parse_date_range,
    validate_excel_file,
    validate_file_size,
    create_temp_file,
    cleanup_temp_file,
    parse_template_json,
    get_firebird_service
)
from app.middleware.rate_limit import limiter
from app.auth import require_auth_if_enabled, require_admin
from app.services.logging_service import logging_service
from app.services.cache_service import (
    CacheService,
    invalidate_transactions_cache,
    invalidate_dashboard_cache
)
import hashlib
import json

settings = get_settings()
cache = CacheService.get_instance()

router = APIRouter(prefix="/api/v1/transactions", tags=["transactions"])


@router.post("/upload", response_model=FileUploadResponse)
@limiter.limit(settings.rate_limit_strict)
async def upload_file(
    request: Request,
    file: UploadFile = File(...),
    provider_id: Optional[int] = Query(None, description="ID провайдера (если не указан, будет попытка автоопределения)"),
    template_id: Optional[int] = Query(None, description="ID шаблона (если не указан, будет попытка автоопределения)"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Загрузка Excel файла и создание транзакций
    Автоматически определяет провайдера и шаблон на основе структуры файла.
    Если автоопределение не удалось, возвращает require_template_selection=true с списком доступных шаблонов.
    """
    logger.info(
        f"Начало загрузки файла: {file.filename}",
        extra={
            "file_name": file.filename,
            "content_type": file.content_type,
            "method": request.method,
            "path": request.url.path
        }
    )
    start_time = datetime.now()
    event_service = UploadEventService(db)
    provider_id = None
    template_id = None
    transactions_total = 0
    created_count = 0
    skipped_count = 0
    
    # Валидация типа файла
    try:
        validate_excel_file(file)
    except HTTPException as e:
        logger.error(
            "Ошибка валидации файла",
            extra={
                "file_name": file.filename,
                "error": str(e.detail),
                "status_code": e.status_code
            }
        )
        raise
    
    # Валидация размера файла из конфигурации
    MAX_FILE_SIZE = settings.max_upload_size
    try:
        content = await file.read()
        file_size = len(content)
        
        logger.debug(
            f"Размер загружаемого файла: {file_size / 1024 / 1024:.2f}MB",
            extra={"file_size_bytes": file_size}
        )
        
        validate_file_size(content, MAX_FILE_SIZE)
    except HTTPException as e:
        logger.error(
            "Ошибка валидации размера файла",
            extra={
                "file_name": file.filename,
                "error": str(e.detail),
                "status_code": e.status_code
            }
        )
        raise
    except Exception as e:
        logger.error(
            "Ошибка при чтении файла",
            extra={
                "file_name": file.filename,
                "error": str(e),
                "error_type": type(e).__name__
            },
            exc_info=True
        )
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при чтении файла: {str(e)}"
        )
    
    # Сохраняем файл во временную директорию
    tmp_file_path = None
    try:
        tmp_file_path = create_temp_file(content, suffix=".xlsx")
        
        match_info = {}
        auto_detected = False
        
        # Если провайдер и шаблон не указаны, пытаемся автоопределить
        if not provider_id or not template_id:
            logger.info("Определение провайдера и шаблона для файла", extra={"file_name": file.filename})
            detected_provider_id, detected_template_id, match_info = detect_provider_and_template(tmp_file_path, db)
            
            # Используем автоопределенные значения, если они не были переданы
            if not provider_id:
                provider_id = detected_provider_id
            if not template_id:
                template_id = detected_template_id
            
            auto_detected = True
            
            # Если провайдер не определен, используем провайдера по умолчанию
            if not provider_id:
                default_provider = db.query(Provider).filter(Provider.code == "RP-GAZPROM").first()
                if default_provider:
                    provider_id = default_provider.id
                    logger.info("Использован провайдер по умолчанию", extra={"provider_id": provider_id})
        
        # Проверяем, удалось ли автоопределение (score >= 30 и template_id определен)
        match_score = match_info.get("score", 0) if match_info else 0
        requires_selection = auto_detected and (match_score < 30 or not template_id)
        
        if requires_selection:
            # Собираем список всех доступных провайдеров и их шаблонов
            providers = db.query(Provider).filter(Provider.is_active == True).all()
            available_templates = []
            
            for provider in providers:
                templates = db.query(ProviderTemplate).filter(
                    ProviderTemplate.provider_id == provider.id,
                    ProviderTemplate.is_active == True
                ).all()
                
                for template in templates:
                    available_templates.append({
                        "template_id": template.id,
                        "template_name": template.name,
                        "provider_id": provider.id,
                        "provider_name": provider.name,
                        "provider_code": provider.code
                    })
            
            logger.info(
                "Автоопределение не удалось, требуется выбор шаблона",
                extra={
                    "file_name": file.filename,
                    "match_score": match_score,
                    "available_templates_count": len(available_templates)
                }
            )
            
            # Возвращаем ответ с требованием выбора шаблона
            return JSONResponse(
                status_code=200,
                content=FileUploadResponse(
                    message="Не удалось автоматически определить шаблон. Пожалуйста, выберите шаблон вручную.",
                    transactions_created=0,
                    transactions_skipped=0,
                    file_name=file.filename,
                    validation_warnings=[],
                    require_template_selection=True,
                    available_templates=available_templates,
                    detected_provider_id=provider_id,
                    detected_template_id=template_id,
                    match_info=match_info
                ).model_dump()
            )
        
        if provider_id:
            logger.info(
                "Провайдер определен",
                extra={
                    "provider_id": provider_id,
                    "template_id": template_id,
                    "match_score": match_info.get("score", 0) if match_info else 0
                }
            )
        
        # Определяем стратегию обработки на основе размера файла
        # Для файлов больше 10MB используем chunked reading
        CHUNKED_THRESHOLD = 10 * 1024 * 1024  # 10MB
        use_chunked = file_size > CHUNKED_THRESHOLD
        chunk_size = 1000 if use_chunked else None
        
        logger.info(
            "Обработка Excel файла",
            extra={
                "file_name": file.filename,
                "file_size_mb": round(file_size / 1024 / 1024, 2),
                "use_chunked": use_chunked,
                "chunk_size": chunk_size
            }
        )
        
        # Используем оптимизированный процессор
        excel_processor = ExcelProcessor(db)
        transactions_data = excel_processor.process_file(
            tmp_file_path,
            file.filename,
            provider_id=provider_id,
            template_id=template_id,
            chunk_size=chunk_size
        )
        
        if not transactions_data:
            logger.warning("В файле не найдено транзакций", extra={"file_name": file.filename})
            raise HTTPException(status_code=400, detail="Не найдено транзакций в файле")
        
        transactions_total = len(transactions_data)
        logger.info(f"Найдено транзакций в файле: {len(transactions_data)}", extra={"transactions_count": len(transactions_data)})
        
        # Проверяем дату закрытия периода загрузки
        period_lock = db.query(UploadPeriodLock).first()
        if period_lock:
            # Проверяем, есть ли транзакции с датами раньше даты закрытия периода
            blocked_transactions = []
            for trans_data in transactions_data:
                trans_date = trans_data.get("transaction_date")
                if trans_date and isinstance(trans_date, datetime):
                    trans_date_only = trans_date.date()
                    if trans_date_only < period_lock.lock_date:
                        blocked_transactions.append(trans_date_only)
            
            if blocked_transactions:
                min_date = min(blocked_transactions)
                raise HTTPException(
                    status_code=400,
                    detail=f"Нельзя загружать транзакции с датами раньше {period_lock.lock_date.strftime('%d.%m.%Y')}. "
                           f"Найдены транзакции с датой {min_date.strftime('%d.%m.%Y')}"
                )
        
        # Создаем транзакции в БД с батчевой обработкой
        logger.info("Создание транзакций в базе данных", extra={"transactions_count": len(transactions_data)})
        
        try:
            batch_processor = TransactionBatchProcessor(db)
            created_count, skipped_count, warnings = batch_processor.create_transactions(transactions_data)
            
            logger.info(
                "Файл успешно обработан",
                extra={
                    "created_count": created_count,
                    "skipped_count": skipped_count,
                    "warnings_count": len(warnings),
                    "file_name": file.filename
                }
            )
        except Exception as e:
            db.rollback()
            logger.error(
                "Ошибка при создании транзакций в БД",
                extra={
                    "file_name": file.filename,
                    "transactions_count": len(transactions_data),
                    "error": str(e)
                },
                exc_info=True
            )
            raise HTTPException(
                status_code=500,
                detail=f"Ошибка при создании транзакций в базе данных: {str(e)}"
            )
        
        message = f"Файл успешно обработан. Создано транзакций: {created_count}"
        if skipped_count > 0:
            message += f", пропущено дубликатов: {skipped_count}"
        if warnings:
            message += f". Обнаружено предупреждений: {len(warnings)}"
            logger.warning("Обнаружены предупреждения при обработке файла", extra={"warnings": warnings})
        
        # Добавляем информацию об определенном провайдере
        if match_info.get("provider_name"):
            message += f"\n\nОпределен провайдер: {match_info['provider_name']}"
            if match_info.get("template_name"):
                message += f" (шаблон: {match_info['template_name']})"
            if match_info.get("score", 0) > 0:
                message += f" (совпадение: {match_info['score']}%)"

        event_service.log_event(
            source_type="manual",
            status="success",
            is_scheduled=False,
            file_name=file.filename,
            provider_id=provider_id,
            template_id=template_id,
            user_id=current_user.id if current_user else None,
            username=current_user.username if current_user else None,
            transactions_total=transactions_total,
            transactions_created=created_count,
            transactions_skipped=skipped_count,
            transactions_failed=0,
            duration_ms=int((datetime.now() - start_time).total_seconds() * 1000),
            message="; ".join(warnings) if warnings else message
        )
        
        # Логируем действие пользователя
        if current_user:
            try:
                logging_service.log_user_action(
                    db=db,
                    user_id=current_user.id,
                    username=current_user.username,
                    action_type="upload",
                    action_description=f"Загружен файл транзакций: {file.filename}",
                    action_category="transaction",
                    entity_type="Transaction",
                    entity_id=None,
                    status="success",
                    extra_data={
                        "file_name": file.filename,
                        "provider_id": provider_id,
                        "template_id": template_id,
                        "transactions_total": transactions_total,
                        "transactions_created": created_count,
                        "transactions_skipped": skipped_count
                    }
                )
            except Exception as e:
                logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
        
        # Инвалидируем кэш при успешной загрузке
        if created_count > 0:
            invalidate_transactions_cache()
            invalidate_dashboard_cache()
            logger.debug("Кэш транзакций и дашборда инвалидирован после загрузки файла")
        
        response_data = FileUploadResponse(
            message=message,
            transactions_created=created_count,
            transactions_skipped=skipped_count,
            file_name=file.filename,
            validation_warnings=warnings,
            require_template_selection=False,
            available_templates=None,
            detected_provider_id=provider_id,
            detected_template_id=template_id,
            match_info=match_info if match_info else None
        )
        
        return JSONResponse(
            status_code=200,
            content=response_data.model_dump()
        )
    except HTTPException as http_exc:
        # Логируем событие неуспешной загрузки
        event_service.log_event(
            source_type="manual",
            status="failed",
            is_scheduled=False,
            file_name=file.filename,
            provider_id=provider_id,
            template_id=template_id,
            user_id=current_user.id if current_user else None,
            username=current_user.username if current_user else None,
            transactions_total=transactions_total,
            transactions_created=created_count,
            transactions_skipped=skipped_count,
            transactions_failed=transactions_total - created_count - skipped_count if transactions_total else 0,
            duration_ms=int((datetime.now() - start_time).total_seconds() * 1000),
            message=str(http_exc.detail)
        )
        raise
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        logger.error(
            f"Ошибка при обработке файла: {file.filename}",
            extra={
                "file_name": file.filename, 
                "error": str(e),
                "error_type": type(e).__name__,
                "traceback": error_traceback
            },
            exc_info=True
        )
        event_service.log_event(
            source_type="manual",
            status="failed",
            is_scheduled=False,
            file_name=file.filename,
            provider_id=provider_id,
            template_id=template_id,
            user_id=current_user.id if current_user else None,
            username=current_user.username if current_user else None,
            transactions_total=transactions_total,
            transactions_created=created_count,
            transactions_skipped=skipped_count,
            transactions_failed=transactions_total - created_count - skipped_count if transactions_total else 0,
            duration_ms=int((datetime.now() - start_time).total_seconds() * 1000),
            message=str(e)
        )
        # Возвращаем более детальную информацию об ошибке
        error_detail = f"Ошибка при обработке файла: {str(e)}"
        if hasattr(e, '__cause__') and e.__cause__:
            error_detail += f" (причина: {str(e.__cause__)})"
        raise HTTPException(status_code=500, detail=error_detail)
    finally:
        # Удаляем временный файл
        if tmp_file_path:
            cleanup_temp_file(tmp_file_path)


@router.post("/load-from-api", response_model=FileUploadResponse)
async def load_from_api(
    template_id: int = Query(..., description="ID шаблона с типом подключения 'api' или 'web'"),
    date_from: Optional[str] = Query(None, description="Начальная дата периода в формате YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="Конечная дата периода в формате YYYY-MM-DD"),
    card_numbers: Optional[str] = Query(None, description="Список номеров карт через запятую (опционально)"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Загрузка транзакций через API или веб-сервис провайдера
    """
    from datetime import date as date_type
    import sys
    
    logger.info("POST /api/v1/transactions/load-from-api ВЫЗВАН (РУЧНАЯ ЗАГРУЗКА)", extra={
        "template_id": template_id,
        "date_from": date_from,
        "date_to": date_to,
        "card_numbers": card_numbers
    })
    
    template = db.query(ProviderTemplate).filter(ProviderTemplate.id == template_id).first()
    
    logger.info("=" * 80)
    logger.info("Начало загрузки данных из API (ручная загрузка)", extra={
        "template_id": template_id,
        "date_from": date_from,
        "date_to": date_to,
        "card_numbers": card_numbers,
        "user_id": current_user.id if current_user else None,
        "username": current_user.username if current_user else None,
        "event_type": "manual_load",
        "event_category": "api"
    })
    logger.info("=" * 80)
    if not template:
        raise HTTPException(status_code=404, detail="Шаблон не найден")
    
    if template.connection_type not in ["api", "web"]:
        raise HTTPException(
            status_code=400,
            detail=f"Шаблон имеет тип подключения '{template.connection_type}', ожидается 'api' или 'web'"
        )
    
    # Парсим даты
    if not date_from or not date_to:
        raise HTTPException(
            status_code=400,
            detail="Необходимо указать date_from и date_to"
        )
    
    try:
        parsed_date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        parsed_date_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError as e:
        raise HTTPException(
            status_code=400,
            detail=f"Неверный формат даты. Используйте формат YYYY-MM-DD. Ошибка: {str(e)}"
        )
    
    # Парсим список карт (если указан)
    card_list = None
    if card_numbers:
        card_list = [card.strip() for card in card_numbers.split(",") if card.strip()]
    
    # Проверяем дату закрытия периода загрузки
    period_lock = db.query(UploadPeriodLock).first()
    if period_lock:
        if parsed_date_from < period_lock.lock_date:
            raise HTTPException(
                status_code=400,
                detail=f"Нельзя загружать транзакции с датами раньше {period_lock.lock_date.strftime('%d.%m.%Y')}. "
                       f"Указана дата начала: {parsed_date_from.strftime('%d.%m.%Y')}"
            )
    
    event_service = UploadEventService(db)
    start_time = datetime.now()
    
    try:
        # Загружаем транзакции через API
        api_service = ApiProviderService(db)
        import asyncio
        transactions_data = await api_service.fetch_transactions(
            template=template,
            date_from=parsed_date_from,
            date_to=parsed_date_to,
            card_numbers=card_list
        )
        
        if not transactions_data:
            duration_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            event_service.log_event(
                source_type="manual",
                status="success",
                is_scheduled=False,
                file_name=f"API: {template.name}",
                provider_id=template.provider_id,
                template_id=template.id,
                user_id=current_user.id if current_user else None,
                username=current_user.username if current_user else None,
                transactions_total=0,
                transactions_created=0,
                transactions_skipped=0,
                transactions_failed=0,
                duration_ms=duration_ms,
                message="Транзакции не найдены за указанный период"
            )
            return FileUploadResponse(
                message="Транзакции не найдены за указанный период",
                transactions_created=0,
                transactions_skipped=0,
                file_name=f"API_{template.name}"
            )
        
        logger.info(f"Загружено транзакций из API: {len(transactions_data)}", extra={
            "template_id": template_id,
            "transactions_count": len(transactions_data)
        })
        
        # Парсим маппинг видов топлива (если указан)
        fuel_type_mapping = None
        try:
            if template.fuel_type_mapping:
                fuel_type_mapping = parse_template_json(template.fuel_type_mapping)
                import sys
                logger.debug(f"✓ Маппинг топлива загружен для шаблона {template.id} (ручная загрузка API): {list(fuel_type_mapping.keys()) if isinstance(fuel_type_mapping, dict) else 'не словарь'}")
                logger.info("Маппинг видов топлива (API, ручная загрузка) загружен", extra={
                    "template_id": template.id,
                    "template_name": template.name,
                    "mapping_keys": list(fuel_type_mapping.keys()) if isinstance(fuel_type_mapping, dict) else None,
                    "mapping": fuel_type_mapping if isinstance(fuel_type_mapping, dict) else None,
                    "event_type": "manual_load",
                    "event_category": "fuel_mapping"
                })
            else:
                fuel_type_mapping = None
                import sys
                logger.warning(f"✗ Маппинг топлива НЕ НАЙДЕН для шаблона {template.id} (ручная загрузка API)")
        except Exception as fuel_map_err:
            import sys
            print(f"✗ ОШИБКА при загрузке маппинга топлива для шаблона {template.id}: {fuel_map_err}", file=sys.stderr, flush=True)
            logger.warning("Не удалось разобрать маппинг видов топлива (ручная загрузка API)", extra={
                "template_id": template.id,
                "template_name": template.name,
                "error": str(fuel_map_err),
                "event_type": "manual_load",
                "event_category": "fuel_mapping"
            })
        
        # Применяем маппинг топлива к транзакциям (если указан)
        if fuel_type_mapping and isinstance(fuel_type_mapping, dict):
            # Функция для сопоставления топлива с маппингом (из auto_load_service)
            def _match_fuel(fuel_name: str, mapping: dict) -> Optional[str]:
                if not fuel_name or not mapping:
                    return None
                norm = str(fuel_name).strip().lower().replace(" ", "").replace("-", "")
                for source_name, target_name in mapping.items():
                    src_norm = str(source_name).strip().lower().replace(" ", "").replace("-", "")
                    if src_norm == norm:
                        return target_name
                    # Проверяем обратное соответствие: если в данных уже "нормализованное" значение, оставляем как есть
                    if target_name:
                        tgt_norm = str(target_name).strip().lower().replace(" ", "").replace("-", "")
                        if tgt_norm == norm:
                            return target_name
                return None
            
            from app import services as app_services
            mapped_count = 0
            for transaction in transactions_data:
                if "product" in transaction:
                    raw_fuel = str(transaction["product"] or "")
                    if raw_fuel:
                        mapped = _match_fuel(raw_fuel, fuel_type_mapping)
                        if mapped:
                            transaction["product"] = mapped
                            mapped_count += 1
                            import sys
                            if mapped_count <= 5:  # Логируем только первые 5 для избежания спама
                                logger.debug(f"  → Маппинг применен (ручная загрузка API): '{raw_fuel}' → '{mapped}'")
                            logger.info("Маппинг топлива применен (API, ручная загрузка)", extra={
                                "template_id": template.id,
                                "template_name": template.name,
                                "raw_fuel": raw_fuel,
                                "mapped_fuel": mapped,
                                "event_type": "manual_load",
                                "event_category": "fuel_mapping"
                            })
            
            if mapped_count > 0:
                import sys
                logger.info(f"✓ Применен маппинг топлива к {mapped_count} транзакциям (ручная загрузка API)")
        
        # Создаем транзакции в БД с батчевой обработкой
        batch_processor = TransactionBatchProcessor(db)
        created_count, skipped_count, warnings = batch_processor.create_transactions(transactions_data)
        
        logger.info(
            "Транзакции из API успешно загружены",
            extra={
                "template_id": template_id,
                "created_count": created_count,
                "skipped_count": skipped_count,
                "warnings_count": len(warnings)
            }
        )
        
        message = f"Успешно загружено транзакций из API: {created_count}"
        if skipped_count > 0:
            message += f", пропущено дубликатов: {skipped_count}"
        
        duration_ms = int((datetime.now() - start_time).total_seconds() * 1000)
        event_service.log_event(
            source_type="manual",
            status="success",
            is_scheduled=False,
            file_name=f"API: {template.name}",
            provider_id=template.provider_id,
            template_id=template.id,
            user_id=current_user.id if current_user else None,
            username=current_user.username if current_user else None,
            transactions_total=len(transactions_data),
            transactions_created=created_count,
            transactions_skipped=skipped_count,
            transactions_failed=0,
            duration_ms=duration_ms,
            message=message
        )
        
        # Инвалидируем кэш при успешной загрузке
        if created_count > 0:
            invalidate_transactions_cache()
            invalidate_dashboard_cache()
            logger.debug("Кэш транзакций и дашборда инвалидирован после загрузки из API")
        
        return FileUploadResponse(
            message=message,
            transactions_created=created_count,
            transactions_skipped=skipped_count,
            file_name=f"API_{template.name}",
            validation_warnings=warnings
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        duration_ms = int((datetime.now() - start_time).total_seconds() * 1000)
        error_message = str(e)
        try:
            event_service.log_event(
                source_type="manual",
                status="failed",
                is_scheduled=False,
                file_name=f"API: {template.name}",
                provider_id=template.provider_id,
                template_id=template.id,
                user_id=current_user.id if current_user else None,
                username=current_user.username if current_user else None,
                transactions_total=0,
                transactions_created=0,
                transactions_skipped=0,
                transactions_failed=0,
                duration_ms=duration_ms,
                message=error_message
            )
        except Exception:
            logger.warning("Не удалось зафиксировать событие загрузки из API", exc_info=True)
        
        logger.error(
            "Ошибка при загрузке транзакций из API",
            extra={
                "template_id": template_id,
                "error": error_message
            },
            exc_info=True
        )
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при загрузке транзакций из API: {error_message}"
        )


@router.post("/load-from-firebird", response_model=FileUploadResponse)
async def load_from_firebird(
    template_id: int = Query(..., description="ID шаблона провайдера с типом подключения firebird"),
    date_from: Optional[str] = Query(None, description="Начальная дата периода в формате YYYY-MM-DD или YYYY-MM-DD HH:MM:SS"),
    date_to: Optional[str] = Query(None, description="Конечная дата периода в формате YYYY-MM-DD или YYYY-MM-DD HH:MM:SS"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Загрузка транзакций из базы данных Firebird за указанный период
    Использует настройки подключения и маппинг полей из шаблона провайдера
    
    Параметры периода:
    - date_from: Начальная дата (включительно). Если не указана, фильтрация не применяется.
    - date_to: Конечная дата (включительно). Если не указана, фильтрация не применяется.
    """
    import sys
    msg = f"\n{'='*80}\nPOST /api/v1/transactions/load-from-firebird ВЫЗВАН (РУЧНАЯ ЗАГРУЗКА)\nTemplate ID: {template_id}\nDate from: {date_from}\nDate to: {date_to}\n{'='*80}\n"
    print(msg, file=sys.stderr, flush=True)
    print(msg, file=sys.stdout, flush=True)
    
    # Проверяем доступность Firebird
    firebird_service_class = get_firebird_service()
    
    logger.info("=" * 80)
    logger.info("Начало загрузки данных из Firebird (ручная загрузка)", extra={
        "template_id": template_id,
        "date_from": date_from,
        "date_to": date_to,
        "user_id": current_user.id if current_user else None,
        "username": current_user.username if current_user else None,
        "event_type": "manual_load",
        "event_category": "firebird"
    })
    logger.info("=" * 80)
    
    # Получаем шаблон
    template = db.query(ProviderTemplate).filter(
        ProviderTemplate.id == template_id,
        ProviderTemplate.is_active == True
    ).first()
    
    if not template:
        raise HTTPException(status_code=404, detail="Шаблон не найден или неактивен")
    
    if template.connection_type != "firebird":
        raise HTTPException(
            status_code=400,
            detail=f"Шаблон имеет тип подключения '{template.connection_type}', ожидается 'firebird'"
        )
    
    if not template.connection_settings:
        raise HTTPException(
            status_code=400,
            detail="В шаблоне не указаны настройки подключения к Firebird"
        )
    
    event_service = UploadEventService(db)
    start_time = datetime.now()
    
    try:
        # Парсим настройки подключения
        connection_settings = parse_template_json(template.connection_settings)
        
        # Парсим маппинг полей
        field_mapping = parse_template_json(template.field_mapping)
        
        # Парсим маппинг видов топлива
        fuel_type_mapping = None
        try:
            if template.fuel_type_mapping:
                fuel_type_mapping = parse_template_json(template.fuel_type_mapping)
                import sys
                print(f"✓ Маппинг топлива загружен для шаблона {template.id} (ручная загрузка)", file=sys.stderr, flush=True)
                if isinstance(fuel_type_mapping, dict):
                    print(f"  Ключи маппинга (исходные значения): {list(fuel_type_mapping.keys())}", file=sys.stderr, flush=True)
                    print(f"  Значения маппинга (нормализованные): {list(fuel_type_mapping.values())}", file=sys.stderr, flush=True)
                    print(f"  Полный маппинг: {fuel_type_mapping}", file=sys.stderr, flush=True)
                else:
                    print(f"  ⚠ Маппинг не является словарем: {type(fuel_type_mapping)}", file=sys.stderr, flush=True)
                logger.info("Маппинг видов топлива (Firebird, ручная загрузка) загружен", extra={
                    "template_id": template.id,
                    "template_name": template.name,
                    "mapping_keys": list(fuel_type_mapping.keys()) if isinstance(fuel_type_mapping, dict) else None,
                    "mapping": fuel_type_mapping if isinstance(fuel_type_mapping, dict) else None,
                    "event_type": "manual_load",
                    "event_category": "fuel_mapping"
                })
            else:
                fuel_type_mapping = None
                import sys
                print(f"✗ Маппинг топлива НЕ НАЙДЕН для шаблона {template.id} (ручная загрузка)", file=sys.stderr, flush=True)
        except Exception as fuel_map_err:
            import sys
            print(f"✗ ОШИБКА при загрузке маппинга топлива для шаблона {template.id}: {fuel_map_err}", file=sys.stderr, flush=True)
            logger.warning("Не удалось разобрать маппинг видов топлива (ручная загрузка)", extra={
                "template_id": template.id,
                "template_name": template.name,
                "error": str(fuel_map_err),
                "event_type": "manual_load",
                "event_category": "fuel_mapping"
            })
        
        # Парсим даты периода, если указаны
        parsed_date_from, parsed_date_to = parse_date_range(date_from, date_to)
        
        # Читаем данные из Firebird
        firebird_service = firebird_service_class(db)
        firebird_data = firebird_service.read_data(
            connection_settings=connection_settings,
            source_table=template.source_table,
            source_query=template.source_query,
            field_mapping=field_mapping,
            date_from=parsed_date_from,
            date_to=parsed_date_to
        )
        
        logger.info("Данные прочитаны из Firebird", extra={
            "rows_count": len(firebird_data),
            "template_id": template_id
        })
        
        # Собираем уникальные значения топлива из базы данных для диагностики
        unique_fuel_values = set()
        for row in firebird_data[:100]:  # Проверяем первые 100 строк
            fuel_val = str(row.get("fuel") or row.get("product") or "").strip()
            if fuel_val:
                unique_fuel_values.add(fuel_val)
        
        import sys
        print(f"  [DEBUG] Уникальные значения топлива из БД (первые 100 строк): {sorted(unique_fuel_values)}", file=sys.stderr, flush=True)
        logger.info("Уникальные значения топлива из базы данных", extra={
            "template_id": template_id,
            "unique_fuel_values": sorted(unique_fuel_values),
            "event_type": "manual_load",
            "event_category": "fuel_mapping"
        })
        
        # Преобразуем данные из Firebird в формат транзакций
        from app import services as app_services
        from decimal import Decimal
        
        transactions_data = []
        warnings = []
        warning_counts = {}  # Счетчик одинаковых предупреждений
        
        # Определяем имена полей из маппинга для более гибкого поиска
        quantity_field_names = ["quantity", "qty"]
        date_field_names = ["date", "transaction_date"]
        
        if field_mapping:
            # Ищем поле количества в маппинге
            for sys_field, db_field in field_mapping.items():
                if sys_field.lower() in ['quantity', 'qty', 'количество']:
                    quantity_field_names.append(db_field)
                if sys_field.lower() in ['date', 'transaction_date', 'дата']:
                    date_field_names.append(db_field)
        
        # Логируем доступные поля из первой строки для отладки
        if firebird_data and len(firebird_data) > 0:
            available_fields = list(firebird_data[0].keys())
            logger.info("Доступные поля в данных Firebird", extra={
                "available_fields": available_fields,
                "quantity_field_names": quantity_field_names,
                "date_field_names": date_field_names,
                "field_mapping": field_mapping
            })
        
        for row_idx, row in enumerate(firebird_data):
            try:
                transaction_data = {}
                
                # Дата транзакции - ищем по всем возможным именам
                date_value = None
                for field_name in date_field_names:
                    date_value = row.get(field_name)
                    if not date_value:
                        # Пробуем регистронезависимый поиск
                        for key, val in row.items():
                            if key.lower() == field_name.lower():
                                date_value = val
                                break
                    if date_value:
                        break
                
                if date_value:
                    if isinstance(date_value, datetime):
                        transaction_data["transaction_date"] = date_value
                    else:
                        parsed_date = app_services.parse_excel_date(date_value)
                        if parsed_date:
                            transaction_data["transaction_date"] = parsed_date
                        else:
                            error_msg = f"Не удалось распарсить дату: {date_value}"
                            if error_msg not in warning_counts:
                                warning_counts[error_msg] = 0
                            warning_counts[error_msg] += 1
                            if warning_counts[error_msg] <= 5:
                                warnings.append(f"Строка {row_idx + 1}: {error_msg}")
                            elif warning_counts[error_msg] == 6:
                                warnings.append(f"... и еще {len(firebird_data) - row_idx} строк с той же ошибкой")
                            continue
                else:
                    available_date_fields = [k for k in row.keys() if 'date' in k.lower() or 'дата' in k.lower()]
                    error_msg = f"Не найдена дата транзакции. Искали в полях: {date_field_names}. Доступные поля с датой: {available_date_fields if available_date_fields else 'нет'}"
                    if row_idx == 0:
                        error_msg += f". Все доступные поля в данных: {list(row.keys())}"
                    
                    if error_msg not in warning_counts:
                        warning_counts[error_msg] = 0
                    warning_counts[error_msg] += 1
                    
                    if warning_counts[error_msg] <= 5:
                        warnings.append(f"Строка {row_idx + 1}: {error_msg}")
                    elif warning_counts[error_msg] == 6:
                        warnings.append(f"... и еще {len(firebird_data) - row_idx} строк с той же ошибкой")
                    continue
                
                # Количество - ищем по всем возможным именам
                qty_value = None
                found_qty_field = None
                for field_name in quantity_field_names:
                    qty_value = row.get(field_name)
                    if not qty_value:
                        # Пробуем регистронезависимый поиск
                        for key, val in row.items():
                            if key.lower() == field_name.lower():
                                qty_value = val
                                found_qty_field = key
                                break
                    if qty_value is not None:
                        found_qty_field = found_qty_field or field_name
                        break
                
                if qty_value is not None:
                    qty_decimal = app_services.convert_to_decimal(qty_value)
                    if qty_decimal is not None:
                        # Инвертируем отрицательные значения количества
                        # Если значение отрицательное, делаем его положительным
                        if qty_decimal < 0:
                            qty_decimal = abs(qty_decimal)
                        transaction_data["quantity"] = qty_decimal
                    else:
                        warnings.append(f"Строка {row_idx + 1}: Не удалось преобразовать количество из поля '{found_qty_field}': {qty_value} (тип: {type(qty_value).__name__})")
                        continue
                else:
                    # Ищем поля, которые могут быть количеством (содержат qty, quantity, amount, количество)
                    possible_qty_fields = [k for k in row.keys() if any(term in k.lower() for term in ['qty', 'quantity', 'amount', 'количество', 'литр', 'литры'])]
                    
                    # Формируем сообщение об ошибке
                    error_msg = f"Не найдено количество. Искали в полях: {quantity_field_names}. Доступные поля, похожие на количество: {possible_qty_fields if possible_qty_fields else 'нет'}"
                    if row_idx == 0:  # Только для первой строки показываем все доступные поля
                        error_msg += f". Все доступные поля в данных: {list(row.keys())}"
                    
                    # Подсчитываем одинаковые предупреждения
                    if error_msg not in warning_counts:
                        warning_counts[error_msg] = 0
                    warning_counts[error_msg] += 1
                    
                    # Показываем предупреждение только первые 5 раз, затем сводку
                    if warning_counts[error_msg] <= 5:
                        warnings.append(f"Строка {row_idx + 1}: {error_msg}")
                    elif warning_counts[error_msg] == 6:
                        warnings.append(f"... и еще {len(firebird_data) - row_idx} строк с той же ошибкой")
                    continue
                
                # Остальные поля
                transaction_data["card_number"] = str(row.get("card") or row.get("card_number") or "").strip()
                transaction_data["vehicle"] = str(row.get("user") or row.get("vehicle") or "").strip()
                kazs_value = str(row.get("kazs") or row.get("azs_number") or "").strip()
                transaction_data["azs_number"] = app_services.extract_azs_number(kazs_value)
                transaction_data["azs_original_name"] = kazs_value  # Сохраняем оригинальное название АЗС
                # Применяем маппинг топлива (если указан)
                raw_fuel = str(row.get("fuel") or row.get("product") or "")
                
                # Логируем исходное значение из базы данных (только первые несколько для отладки)
                if row_idx < 5 and raw_fuel:
                    import sys
                    print(f"  [DEBUG] Исходное значение топлива из БД (строка {row_idx + 1}): '{raw_fuel}'", file=sys.stderr, flush=True)
                
                # Функция для сопоставления топлива с маппингом (из auto_load_service)
                def _match_fuel(fuel_name: str, mapping: dict) -> Optional[str]:
                    if not fuel_name or not mapping:
                        return None
                    norm = str(fuel_name).strip().lower().replace(" ", "").replace("-", "")
                    for source_name, target_name in mapping.items():
                        src_norm = str(source_name).strip().lower().replace(" ", "").replace("-", "")
                        if src_norm == norm:
                            return target_name
                        # Проверяем обратное соответствие: если в данных уже "нормализованное" значение, оставляем как есть
                        if target_name:
                            tgt_norm = str(target_name).strip().lower().replace(" ", "").replace("-", "")
                            if tgt_norm == norm:
                                return target_name
                    return None
                
                normalized_fuel = raw_fuel
                mapping_applied = False
                if raw_fuel and fuel_type_mapping:
                    # Логируем маппинг для отладки (только первые несколько)
                    if row_idx < 5:
                        import sys
                        print(f"  [DEBUG] Пробуем применить маппинг для '{raw_fuel}'", file=sys.stderr, flush=True)
                        print(f"  [DEBUG] Доступные ключи маппинга: {list(fuel_type_mapping.keys())}", file=sys.stderr, flush=True)
                    
                    mapped = _match_fuel(raw_fuel, fuel_type_mapping)
                    if mapped:
                        normalized_fuel = mapped
                        mapping_applied = True
                        import sys
                        # Логируем только если значение изменилось или первые 10 для отладки
                        if normalized_fuel != raw_fuel or row_idx < 10:
                            print(f"  → Маппинг применен (ручная загрузка): '{raw_fuel}' → '{normalized_fuel}'", file=sys.stderr, flush=True)
                        logger.info("Маппинг топлива применен (Firebird, ручная загрузка)", extra={
                            "template_id": template.id,
                            "template_name": template.name,
                            "raw_fuel": raw_fuel,
                            "mapped_fuel": normalized_fuel,
                            "event_type": "manual_load",
                            "event_category": "fuel_mapping"
                        })
                    else:
                        import sys
                        # Логируем только первые несколько для отладки
                        if row_idx < 5:
                            print(f"  → Маппинг НЕ НАЙДЕН для '{raw_fuel}' (ручная загрузка), используем нормализацию", file=sys.stderr, flush=True)
                
                # ВАЖНО: Если маппинг применился, НЕ вызываем normalize_fuel,
                # чтобы избежать обратной нормализации (например, "ДТ" -> "Дизельное топливо")
                # normalize_fuel содержит логику, которая преобразует "ДТ" обратно в "Дизельное топливо"
                if not mapping_applied and normalized_fuel == raw_fuel:
                    # Только если маппинг не сработал, применяем стандартную нормализацию
                    normalized_fuel_before = normalized_fuel
                    normalized_fuel = app_services.normalize_fuel(raw_fuel)
                    # Логируем, если нормализация изменила значение
                    if normalized_fuel != normalized_fuel_before and row_idx < 5:
                        import sys
                        print(f"  [DEBUG] normalize_fuel изменил значение: '{normalized_fuel_before}' -> '{normalized_fuel}'", file=sys.stderr, flush=True)
                
                # Логируем финальное значение перед сохранением в transaction_data (только первые несколько)
                if row_idx < 5:
                    import sys
                    print(f"  [DEBUG] Финальное значение product перед сохранением в transaction_data: '{normalized_fuel}'", file=sys.stderr, flush=True)
                
                transaction_data["product"] = normalized_fuel
                transaction_data["operation_type"] = "Покупка"
                transaction_data["currency"] = "RUB"
                transaction_data["exchange_rate"] = Decimal("1")
                transaction_data["source_file"] = f"Firebird: {template.name}"
                transaction_data["organization"] = str(row.get("organization") or row.get("org") or "").strip()
                transaction_data["provider_id"] = template.provider_id
                
                # Дополнительные поля
                if row.get("supplier"):
                    transaction_data["supplier"] = str(row["supplier"]).strip()
                if row.get("region"):
                    transaction_data["region"] = str(row["region"]).strip()
                if row.get("settlement"):
                    transaction_data["settlement"] = str(row["settlement"]).strip()
                if row.get("location"):
                    transaction_data["location"] = str(row["location"]).strip()
                
                transactions_data.append(transaction_data)
                
            except Exception as e:
                logger.warning("Ошибка преобразования строки данных из Firebird", extra={
                    "error": str(e),
                    "row": str(row)[:200]
                })
                warnings.append(f"Ошибка обработки строки: {str(e)}")
                continue
        
        if not transactions_data:
            # Формируем детальное сообщение об ошибке
            error_detail = "Не удалось преобразовать данные из Firebird в транзакции.\n"
            error_detail += f"Искали поле количества в: {quantity_field_names}\n"
            error_detail += f"Искали поле даты в: {date_field_names}\n"
            if firebird_data and len(firebird_data) > 0:
                error_detail += f"Доступные поля в данных: {list(firebird_data[0].keys())}\n"
            if field_mapping:
                error_detail += f"Текущий маппинг полей: {field_mapping}\n"
            error_detail += "Проверьте маппинг полей в шаблоне и убедитесь, что поле количества правильно сопоставлено."
            
            duration_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            try:
                event_service.log_event(
                    source_type="manual",
                    status="failed",
                    is_scheduled=False,
                    file_name=f"Firebird: {template.name}",
                    provider_id=template.provider_id,
                    template_id=template.id,
                    user_id=current_user.id if current_user else None,
                    username=current_user.username if current_user else None,
                    transactions_total=len(firebird_data) if firebird_data else 0,
                    transactions_created=0,
                    transactions_skipped=0,
                    transactions_failed=len(firebird_data) if firebird_data else 0,
                    duration_ms=duration_ms,
                    message=error_detail
                )
            except Exception:
                logger.warning("Не удалось зафиксировать событие загрузки из Firebird", exc_info=True)
            
            raise HTTPException(
                status_code=400,
                detail=error_detail
            )
        
        # Обрабатываем транзакции батчами
        batch_processor = TransactionBatchProcessor(db)
        created_count, skipped_count = batch_processor.process_transactions_batch(
            transactions_data,
            provider_id=template.provider_id
        )
        
        message = f"Загружено {created_count} транзакций из Firebird"
        if skipped_count > 0:
            message += f" (пропущено: {skipped_count})"
        
        duration_ms = int((datetime.now() - start_time).total_seconds() * 1000)
        try:
            event_service.log_event(
                source_type="manual",
                status="success",
                is_scheduled=False,
                file_name=f"Firebird: {template.name}",
                provider_id=template.provider_id,
                template_id=template.id,
                user_id=current_user.id if current_user else None,
                username=current_user.username if current_user else None,
                transactions_total=len(transactions_data),
                transactions_created=created_count,
                transactions_skipped=skipped_count,
                transactions_failed=0,
                duration_ms=duration_ms,
                message=message
            )
        except Exception:
            logger.warning("Не удалось зафиксировать событие загрузки из Firebird", exc_info=True)
        
        logger.info("Загрузка данных из Firebird завершена", extra={
            "created": created_count,
            "skipped": skipped_count,
            "template_id": template_id
        })
        
        # Инвалидируем кэш при успешной загрузке
        if created_count > 0:
            invalidate_transactions_cache()
            invalidate_dashboard_cache()
            logger.debug("Кэш транзакций и дашборда инвалидирован после загрузки из Firebird")
        
        return FileUploadResponse(
            message=message,
            transactions_created=created_count,
            transactions_skipped=skipped_count,
            file_name=f"Firebird: {template.name}",
            validation_warnings=warnings[:10]
        )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        duration_ms = int((datetime.now() - start_time).total_seconds() * 1000)
        error_detail = f"Ошибка при загрузке данных из Firebird: {str(e)}"
        if hasattr(e, '__cause__') and e.__cause__:
            error_detail += f" (причина: {str(e.__cause__)})"
        
        try:
            event_service.log_event(
                source_type="manual",
                status="failed",
                is_scheduled=False,
                file_name=f"Firebird: {template.name}",
                provider_id=template.provider_id if template else None,
                template_id=template_id,
                user_id=current_user.id if current_user else None,
                username=current_user.username if current_user else None,
                transactions_total=0,
                transactions_created=0,
                transactions_skipped=0,
                transactions_failed=0,
                duration_ms=duration_ms,
                message=error_detail
            )
        except Exception:
            logger.warning("Не удалось зафиксировать событие загрузки из Firebird", exc_info=True)
        
        logger.error(
            "Ошибка при загрузке данных из Firebird",
            extra={
                "template_id": template_id,
                "error": str(e),
                "error_type": type(e).__name__,
                "traceback": error_traceback
            },
            exc_info=True
        )
        raise HTTPException(status_code=500, detail=error_detail)


@router.get("", response_model=TransactionListResponse)
async def get_transactions(
    skip: int = Query(0, ge=0, description="Количество пропущенных записей"),
    limit: int = Query(100, ge=1, le=1000, description="Количество записей"),
    card_number: Optional[str] = Query(None, description="Фильтр по номеру карты"),
    azs_number: Optional[str] = Query(None, description="Фильтр по номеру АЗС"),
    product: Optional[str] = Query(None, description="Фильтр по товару"),
    provider_id: Optional[int] = Query(None, description="Фильтр по ID провайдера"),
    date_from: Optional[str] = Query(None, description="Начальная дата периода в формате YYYY-MM-DD или YYYY-MM-DD HH:MM:SS"),
    date_to: Optional[str] = Query(None, description="Конечная дата периода в формате YYYY-MM-DD или YYYY-MM-DD HH:MM:SS"),
    sort_by: Optional[str] = Query("transaction_date", description="Поле для сортировки"),
    sort_order: Optional[str] = Query("desc", regex="^(asc|desc)$", description="Направление сортировки"),
    db: Session = Depends(get_db),
    _: None = Depends(require_auth_if_enabled)
):
    """
    Получение списка транзакций с пагинацией, фильтрацией и сортировкой
    
    Параметры периода:
    - date_from: Начальная дата (включительно). Если не указана, фильтрация не применяется.
    - date_to: Конечная дата (включительно). Если не указана, фильтрация не применяется.
    """
    try:
        # Парсим даты периода, если указаны
        parsed_date_from, parsed_date_to = parse_date_range(date_from, date_to)
    except Exception as e:
        logger.error(
            "Ошибка при парсинге дат периода",
            extra={
                "date_from": date_from,
                "date_to": date_to,
                "error": str(e),
                "error_type": type(e).__name__
            },
            exc_info=True
        )
        raise HTTPException(
            status_code=400,
            detail=f"Ошибка при парсинге дат периода: {str(e)}"
        )
    
    logger.debug(
        "Запрос списка транзакций",
        extra={
            "skip": skip,
            "limit": limit,
            "card_number": card_number,
            "azs_number": azs_number,
            "product": product,
            "provider_id": provider_id,
            "date_from": parsed_date_from.isoformat() if parsed_date_from else None,
            "date_to": parsed_date_to.isoformat() if parsed_date_to else None,
            "sort_by": sort_by,
            "sort_order": sort_order
        }
    )
    
    # Создаем ключ кэша на основе всех параметров
    cache_key_data = {
        "skip": skip,
        "limit": limit,
        "card_number": card_number,
        "azs_number": azs_number,
        "product": product,
        "provider_id": provider_id,
        "date_from": parsed_date_from.isoformat() if parsed_date_from else None,
        "date_to": parsed_date_to.isoformat() if parsed_date_to else None,
        "sort_by": sort_by,
        "sort_order": sort_order
    }
    cache_key = hashlib.md5(json.dumps(cache_key_data, sort_keys=True).encode()).hexdigest()
    cache_key_full = f"transactions:list:{cache_key}"
    
    # Пробуем получить из кэша (TTL 2 минуты для списков)
    cached_result = cache.get(cache_key_full, prefix="")
    if cached_result is not None:
        logger.debug("Cache hit для списка транзакций", extra={"cache_key": cache_key})
        return TransactionListResponse(**cached_result)
    
    # Используем сервисный слой
    try:
        transaction_service = TransactionService(db)
        result_items, total = transaction_service.get_transactions(
            skip=skip,
            limit=limit,
            card_number=card_number,
            azs_number=azs_number,
            product=product,
            provider_id=provider_id,
            date_from=parsed_date_from,
            date_to=parsed_date_to,
            sort_by=sort_by,
            sort_order=sort_order
        )
        
        logger.info(
            "Список транзакций успешно загружен",
            extra={
                "total": total,
                "returned": len(result_items),
                "skip": skip,
                "limit": limit
            }
        )
        
        result = TransactionListResponse(
            total=total,
            items=result_items
        )
        
        # Кэшируем результат (2 минуты)
        cache.set(
            cache_key_full,
            {"total": result.total, "items": [item.model_dump() for item in result.items]},
            ttl=120,
            prefix=""
        )
        logger.debug("Cache miss, сохранено в кэш", extra={"cache_key": cache_key})
        
        return result
    except Exception as e:
        logger.error(
            "Ошибка при получении списка транзакций",
            extra={
                "skip": skip,
                "limit": limit,
                "error": str(e),
                "error_type": type(e).__name__
            },
            exc_info=True
        )
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при получении списка транзакций: {str(e)}"
        )


@router.delete("/clear")
async def clear_all_transactions(
    confirm: Optional[str] = Query(None, description="Подтверждение удаления всех транзакций"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_admin)
):
    """
    Очистка всех транзакций из базы данных
    """
    # Проверяем параметр confirm (может быть строкой "true" или булевым значением)
    confirm_bool = confirm and confirm.lower() in ("true", "1", "yes")
    
    if not confirm_bool:
        raise HTTPException(
            status_code=400, 
            detail="Для очистки базы данных необходимо установить параметр confirm=true"
        )
    
    try:
        transaction_service = TransactionService(db)
        total_count = transaction_service.clear_all_transactions()
        
        # Логируем действие пользователя
        if current_user:
            try:
                logging_service.log_user_action(
                    db=db,
                    user_id=current_user.id,
                    username=current_user.username,
                    action_type="clear",
                    action_description=f"Очищены все транзакции ({total_count} записей)",
                    action_category="transaction",
                    entity_type="Transaction",
                    entity_id=None,
                    status="success",
                    extra_data={"deleted_count": total_count}
                )
            except Exception as e:
                logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
        
        # Инвалидируем кэш после очистки
        invalidate_transactions_cache()
        invalidate_dashboard_cache()
        logger.debug("Кэш транзакций и дашборда инвалидирован после очистки всех транзакций")
        
        return {
            "message": f"База данных успешно очищена",
            "deleted_count": total_count
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Ошибка при очистке базы данных", extra={"error": str(e)}, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ошибка при очистке базы данных: {str(e)}")


@router.delete("/clear-by-provider")
async def clear_transactions_by_provider(
    provider_id: int = Query(..., description="ID провайдера"),
    date_from: Optional[str] = Query(None, description="Начальная дата периода в формате YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="Конечная дата периода в формате YYYY-MM-DD"),
    confirm: Optional[str] = Query(None, description="Подтверждение удаления транзакций"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_admin)
):
    """
    Очистка транзакций по конкретному провайдеру за все время или за определенный период
    Проверяет заблокированный период перед удалением
    """
    from datetime import date as date_type
    
    # Проверяем параметр confirm
    confirm_bool = confirm and confirm.lower() in ("true", "1", "yes")
    
    if not confirm_bool:
        raise HTTPException(
            status_code=400,
            detail="Для очистки транзакций необходимо установить параметр confirm=true"
        )
    
    # Парсим даты периода, если указаны
    parsed_date_from = None
    parsed_date_to = None
    
    if date_from:
        try:
            parsed_date_from = datetime.strptime(date_from, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Неверный формат начальной даты. Используйте формат YYYY-MM-DD"
            )
    
    if date_to:
        try:
            parsed_date_to = datetime.strptime(date_to, "%Y-%m-%d")
            # Устанавливаем время на конец дня для включения всех транзакций за этот день
            parsed_date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Неверный формат конечной даты. Используйте формат YYYY-MM-DD"
            )
    
    # Проверяем, что date_from <= date_to, если оба указаны
    if parsed_date_from and parsed_date_to and parsed_date_from > parsed_date_to:
        raise HTTPException(
            status_code=400,
            detail="Начальная дата не может быть больше конечной даты"
        )
    
    start_time = datetime.now()
    event_service = UploadEventService(db)
    
    try:
        transaction_service = TransactionService(db)
        result = transaction_service.clear_transactions_by_provider(
            provider_id=provider_id,
            date_from=parsed_date_from,
            date_to=parsed_date_to
        )
        
        # Получаем информацию о провайдере для логирования
        provider = db.query(Provider).filter(Provider.id == provider_id).first()
        provider_name = provider.name if provider else f"ID {provider_id}"
        
        # Формируем сообщение для логирования
        period_info = ""
        if parsed_date_from or parsed_date_to:
            if parsed_date_from:
                period_info += f"с {parsed_date_from.strftime('%d.%m.%Y')}"
            if parsed_date_to:
                if period_info:
                    period_info += f" по {parsed_date_to.strftime('%d.%m.%Y')}"
                else:
                    period_info += f"по {parsed_date_to.strftime('%d.%m.%Y')}"
        else:
            period_info = "за все время"
        
        cleanup_message = (
            f"Очистка транзакций провайдера '{provider_name}' {period_info}. "
            f"Удалено транзакций: {result['deleted_count']}"
        )
        
        # Логируем операцию очистки
        event_service.log_event(
            source_type="cleanup",
            status="success",
            is_scheduled=False,
            file_name=f"Очистка провайдера {provider_name}",
            provider_id=provider_id,
            template_id=None,
            user_id=current_user.id if current_user else None,
            username=current_user.username if current_user else None,
            transactions_total=result['deleted_count'],
            transactions_created=0,
            transactions_skipped=0,
            transactions_failed=0,
            duration_ms=int((datetime.now() - start_time).total_seconds() * 1000),
            message=cleanup_message
        )
        
        logger.info(
            "Транзакции провайдера успешно очищены",
            extra={
                "provider_id": provider_id,
                "provider_name": provider_name,
                "deleted_count": result['deleted_count'],
                "date_from": parsed_date_from.isoformat() if parsed_date_from else None,
                "date_to": parsed_date_to.isoformat() if parsed_date_to else None,
                "user_id": current_user.id if current_user else None,
                "username": current_user.username if current_user else None
            }
        )
        
        # Логируем действие пользователя
        if current_user:
            try:
                logging_service.log_user_action(
                    db=db,
                    user_id=current_user.id,
                    username=current_user.username,
                    action_type="clear",
                    action_description=f"Очищены транзакции провайдера '{provider_name}' {period_info} ({result['deleted_count']} записей)",
                    action_category="transaction",
                    entity_type="Provider",
                    entity_id=provider_id,
                    status="success",
                    extra_data={
                        "provider_id": provider_id,
                        "provider_name": provider_name,
                        "deleted_count": result['deleted_count'],
                        "date_from": parsed_date_from.isoformat() if parsed_date_from else None,
                        "date_to": parsed_date_to.isoformat() if parsed_date_to else None
                    }
                )
            except Exception as e:
                logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
        
        # Инвалидируем кэш после очистки
        invalidate_transactions_cache()
        invalidate_dashboard_cache()
        logger.debug("Кэш транзакций и дашборда инвалидирован после очистки транзакций провайдера")
        
        return {
            "message": result['message'],
            "deleted_count": result['deleted_count']
        }
        
    except ValueError as e:
        db.rollback()
        error_message = str(e)
        
        # Логируем неуспешную операцию
        provider = db.query(Provider).filter(Provider.id == provider_id).first()
        provider_name = provider.name if provider else f"ID {provider_id}"
        
        event_service.log_event(
            source_type="cleanup",
            status="failed",
            is_scheduled=False,
            file_name=f"Очистка провайдера {provider_name}",
            provider_id=provider_id,
            template_id=None,
            user_id=current_user.id if current_user else None,
            username=current_user.username if current_user else None,
            transactions_total=0,
            transactions_created=0,
            transactions_skipped=0,
            transactions_failed=0,
            duration_ms=int((datetime.now() - start_time).total_seconds() * 1000),
            message=f"Ошибка: {error_message}"
        )
        
        logger.error(
            "Ошибка при очистке транзакций провайдера",
            extra={
                "provider_id": provider_id,
                "error": error_message,
                "user_id": current_user.id if current_user else None
            },
            exc_info=True
        )
        
        raise HTTPException(status_code=400, detail=error_message)
        
    except Exception as e:
        db.rollback()
        error_message = f"Ошибка при очистке транзакций провайдера: {str(e)}"
        
        # Логируем неуспешную операцию
        provider = db.query(Provider).filter(Provider.id == provider_id).first()
        provider_name = provider.name if provider else f"ID {provider_id}"
        
        event_service.log_event(
            source_type="cleanup",
            status="failed",
            is_scheduled=False,
            file_name=f"Очистка провайдера {provider_name}",
            provider_id=provider_id,
            template_id=None,
            user_id=current_user.id if current_user else None,
            username=current_user.username if current_user else None,
            transactions_total=0,
            transactions_created=0,
            transactions_skipped=0,
            transactions_failed=0,
            duration_ms=int((datetime.now() - start_time).total_seconds() * 1000),
            message=error_message
        )
        
        logger.error(
            "Ошибка при очистке транзакций провайдера",
            extra={
                "provider_id": provider_id,
                "error": str(e),
                "user_id": current_user.id if current_user else None
            },
            exc_info=True
        )
        
        raise HTTPException(status_code=500, detail=error_message)


@router.get("/{transaction_id}", response_model=TransactionResponse)
async def get_transaction(
    transaction_id: int,
    db: Session = Depends(get_db)
):
    """
    Получение транзакции по ID
    """
    transaction_service = TransactionService(db)
    transaction = transaction_service.get_transaction(transaction_id)
    if not transaction:
        raise HTTPException(status_code=404, detail="Транзакция не найдена")
    return transaction


@router.delete("/{transaction_id}")
async def delete_transaction(
    transaction_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Удаление транзакции по ID
    """
    transaction_service = TransactionService(db)
    success = transaction_service.delete_transaction(transaction_id)
    if not success:
        raise HTTPException(status_code=404, detail="Транзакция не найдена")
    
    # Логируем действие пользователя
    if current_user:
        try:
            logging_service.log_user_action(
                db=db,
                user_id=current_user.id,
                username=current_user.username,
                action_type="delete",
                action_description=f"Удалена транзакция ID: {transaction_id}",
                action_category="transaction",
                entity_type="Transaction",
                entity_id=transaction_id,
                status="success"
            )
        except Exception as e:
            logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
    
    # Инвалидируем кэш транзакций и дашборда
    invalidate_transactions_cache()
    invalidate_dashboard_cache()
    logger.debug("Кэш транзакций и дашборда инвалидирован после удаления транзакции")
    
    return {"message": "Транзакция успешно удалена"}


@router.get("/stats/summary")
async def get_stats_summary(
    provider_id: Optional[int] = Query(None, description="Фильтр по ID провайдера"),
    db: Session = Depends(get_db)
):
    """
    Получение статистики по транзакциям
    """
    transaction_service = TransactionService(db)
    stats = transaction_service.get_stats_summary(provider_id=provider_id)
    
    logger.debug("Статистика по транзакциям загружена", extra={"total_count": stats["total_transactions"]})
    
    return stats


@router.post("/check-match")
async def check_file_match(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Проверка соответствия файла шаблону провайдера перед загрузкой
    Возвращает информацию о совпадении и список доступных шаблонов, если требуется выбор
    """
    validate_excel_file(file)
    
    tmp_file_path = None
    try:
        content = await file.read()
        tmp_file_path = create_temp_file(content, suffix=".xlsx")
        
        # Определяем провайдера и шаблон
        provider_id, template_id, match_info = detect_provider_and_template(tmp_file_path, db)
        
        # Проверяем, требуется ли выбор шаблона
        match_score = match_info.get("score", 0) if match_info else 0
        requires_selection = match_score < 30 or not template_id
        
        available_templates = []
        if requires_selection:
            # Собираем список всех доступных провайдеров и их шаблонов
            providers = db.query(Provider).filter(Provider.is_active == True).all()
            
            for provider in providers:
                templates = db.query(ProviderTemplate).filter(
                    ProviderTemplate.provider_id == provider.id,
                    ProviderTemplate.is_active == True
                ).all()
                
                for template in templates:
                    available_templates.append({
                        "template_id": template.id,
                        "template_name": template.name,
                        "provider_id": provider.id,
                        "provider_name": provider.name,
                        "provider_code": provider.code
                    })
        
        logger.debug("Проверка соответствия файла завершена", extra={
            "match_info": match_info,
            "requires_selection": requires_selection,
            "available_templates_count": len(available_templates)
        })
        
        return {
            "provider_id": provider_id,
            "template_id": template_id,
            "match_info": match_info,
            "is_match": match_score >= 30 and template_id is not None,
            "require_template_selection": requires_selection,
            "available_templates": available_templates if requires_selection else None
        }
    finally:
        if tmp_file_path:
            cleanup_temp_file(tmp_file_path)


@router.get("/export")
async def export_transactions(
    card_number: Optional[str] = Query(None, description="Фильтр по номеру карты"),
    azs_number: Optional[str] = Query(None, description="Фильтр по номеру АЗС"),
    product: Optional[str] = Query(None, description="Фильтр по товару"),
    provider_id: Optional[int] = Query(None, description="Фильтр по ID провайдера"),
    date_from: Optional[str] = Query(None, description="Начальная дата периода в формате YYYY-MM-DD или YYYY-MM-DD HH:MM:SS"),
    date_to: Optional[str] = Query(None, description="Конечная дата периода в формате YYYY-MM-DD или YYYY-MM-DD HH:MM:SS"),
    format: str = Query("xlsx", regex="^(xlsx|csv)$", description="Формат экспорта"),
    db: Session = Depends(get_db)
):
    """
    Экспорт транзакций в Excel или CSV файл
    
    Поддерживает те же фильтры, что и GET /api/v1/transactions
    """
    # Парсим даты периода, если указаны
    parsed_date_from, parsed_date_to = parse_date_range(date_from, date_to)
    
    logger.info(
        "Начало экспорта транзакций",
        extra={
            "card_number": card_number,
            "azs_number": azs_number,
            "product": product,
            "provider_id": provider_id,
            "date_from": parsed_date_from.isoformat() if parsed_date_from else None,
            "date_to": parsed_date_to.isoformat() if parsed_date_to else None,
            "format": format
        }
    )
    
    # Получаем все транзакции с применением фильтров (без пагинации)
    transaction_service = TransactionService(db)
    transactions, total = transaction_service.get_transactions(
        skip=0,
        limit=100000,  # Большой лимит для экспорта всех данных
        card_number=card_number,
        azs_number=azs_number,
        product=product,
        provider_id=provider_id,
        date_from=parsed_date_from,
        date_to=parsed_date_to,
        sort_by="transaction_date",
        sort_order="desc"
    )
    
    if not transactions:
        raise HTTPException(status_code=404, detail="Нет транзакций для экспорта")
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции ГСМ"
    
    # Заголовки
    headers = [
        "ID",
        "Дата и время",
        "№ карты",
        "Закреплена за",
        "Номер АЗС",
        "Товар / услуга",
        "Тип операции",
        "Количество",
        "Валюта",
        "Цена",
        "Сумма",
        "Провайдер",
        "Организация",
        "Источник"
    ]
    
    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Записываем данные
    for row_num, trans in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1, value=trans.id)
        ws.cell(row=row_num, column=2, value=trans.transaction_date.strftime("%d.%m.%Y %H:%M") if trans.transaction_date else "")
        ws.cell(row=row_num, column=3, value=trans.card_number or "")
        ws.cell(row=row_num, column=4, value=trans.vehicle_display_name or trans.vehicle or "")
        ws.cell(row=row_num, column=5, value=trans.azs_number or "")
        ws.cell(row=row_num, column=6, value=trans.product or "")
        ws.cell(row=row_num, column=7, value=trans.operation_type or "Покупка")
        ws.cell(row=row_num, column=8, value=float(trans.quantity) if trans.quantity else "")
        ws.cell(row=row_num, column=9, value=trans.currency or "RUB")
        ws.cell(row=row_num, column=10, value=float(trans.price) if trans.price else "")
        ws.cell(row=row_num, column=11, value=float(trans.amount) if trans.amount else "")
        ws.cell(row=row_num, column=12, value=trans.get("provider_name") or trans.get("supplier") or "")
        ws.cell(row=row_num, column=13, value=trans.get("organization") or "")
        ws.cell(row=row_num, column=14, value=trans.get("source_file") or "")
    
    # Автоматическая ширина колонок
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Сохраняем в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info(
        "Экспорт транзакций завершен",
        extra={
            "total_transactions": total,
            "format": format
        }
    )
    
    # Определяем имя файла
    from datetime import date
    filename = f"transactions_export_{date.today().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
