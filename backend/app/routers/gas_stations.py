"""
Роутер для работы с автозаправочными станциями (АЗС)
"""
from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List, Dict
from datetime import datetime, date
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.database import get_db
from app.logger import logger
from app.models import GasStation, User, Provider
from app.schemas import GasStationResponse, GasStationUpdate, GasStationListResponse
from app.services.gas_station_service import GasStationService
from app.auth import require_auth_if_enabled, require_admin
from app.services.logging_service import logging_service
from app.utils import validate_excel_file, create_temp_file, cleanup_temp_file
from app.middleware.rate_limit import limiter
from app.config import get_settings
from fastapi import Request

router = APIRouter(prefix="/api/v1/gas-stations", tags=["gas-stations"])
settings = get_settings()


@router.get("", response_model=GasStationListResponse)
async def get_gas_stations(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    is_validated: Optional[str] = Query(None, description="Фильтр по статусу валидации: pending, valid, invalid"),
    provider_id: Optional[int] = Query(None, description="Фильтр по ID провайдера"),
    search: Optional[str] = Query(None, description="Поиск по названию, номеру АЗС, местоположению, региону, населенному пункту"),
    sort_by: Optional[str] = Query(None, description="Поле для сортировки: original_name, name, azs_number, location, region, settlement, is_validated, created_at"),
    sort_order: Optional[str] = Query('asc', description="Направление сортировки: asc, desc"),
    db: Session = Depends(get_db)
):
    """
    Получение списка автозаправочных станций
    
    Фильтры:
    - is_validated: pending (требуют проверки), valid (валидные), invalid (с ошибками)
    - provider_id: Фильтр по ID провайдера
    - search: Поиск по названию, номеру АЗС, местоположению, региону, населенному пункту
    - sort_by: Поле для сортировки
    - sort_order: Направление сортировки (asc, desc)
    """
    gas_station_service = GasStationService(db)
    gas_stations, total = gas_station_service.get_gas_stations(
        skip=skip,
        limit=limit,
        is_validated=is_validated,
        provider_id=provider_id,
        search=search,
        sort_by=sort_by,
        sort_order=sort_order
    )
    
    logger.debug("Список АЗС загружен", extra={"total": total, "returned": len(gas_stations), "sort_by": sort_by, "sort_order": sort_order})
    
    return GasStationListResponse(total=total, items=gas_stations)


@router.get("/stats")
async def get_gas_stations_stats(
    db: Session = Depends(get_db)
):
    """
    Получение статистики по АЗС для дашборда
    """
    gas_station_service = GasStationService(db)
    stats = gas_station_service.get_stats_summary()
    
    logger.debug("Статистика АЗС загружена", extra=stats)
    
    return stats


@router.get("/export")
async def export_gas_stations(
    is_validated: Optional[str] = Query(None, description="Фильтр по статусу валидации: pending, valid, invalid"),
    provider_id: Optional[int] = Query(None, description="Фильтр по ID провайдера"),
    search: Optional[str] = Query(None, description="Поиск по названию, номеру АЗС, местоположению"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Экспорт АЗС в Excel файл
    
    Поддерживает те же фильтры, что и GET /api/v1/gas-stations
    """
    logger.info(
        "Начало экспорта АЗС",
        extra={
            "is_validated": is_validated,
            "provider_id": provider_id,
            "search": search
        }
    )
    
    gas_station_service = GasStationService(db)
    gas_stations, total = gas_station_service.get_gas_stations(
        skip=0,
        limit=100000,  # Большой лимит для экспорта всех данных
        is_validated=is_validated,
        provider_id=provider_id,
        search=search
    )
    
    if not gas_stations:
        raise HTTPException(status_code=404, detail="Нет АЗС для экспорта")
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "АЗС"
    
    # Заголовки
    headers = [
        "ID", "Исходное наименование", "Наименование", "Номер АЗС",
        "Местоположение", "Регион", "Населенный пункт",
        "Широта", "Долгота", "Провайдер", "Статус валидации", "Ошибки валидации"
    ]
    
    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Загружаем провайдеров для отображения названий
    provider_ids = [gs.provider_id for gs in gas_stations if gs.provider_id]
    providers_dict = {}
    if provider_ids:
        providers = db.query(Provider).filter(Provider.id.in_(list(set(provider_ids)))).all()
        providers_dict = {p.id: p.name for p in providers}
    
    # Записываем данные
    for row_num, gs in enumerate(gas_stations, 2):
        ws.cell(row=row_num, column=1, value=gs.id)
        ws.cell(row=row_num, column=2, value=gs.original_name or "")
        ws.cell(row=row_num, column=3, value=gs.name or "")
        ws.cell(row=row_num, column=4, value=gs.azs_number or "")
        ws.cell(row=row_num, column=5, value=gs.location or "")
        ws.cell(row=row_num, column=6, value=gs.region or "")
        ws.cell(row=row_num, column=7, value=gs.settlement or "")
        ws.cell(row=row_num, column=8, value=float(gs.latitude) if gs.latitude else "")
        ws.cell(row=row_num, column=9, value=float(gs.longitude) if gs.longitude else "")
        ws.cell(row=row_num, column=10, value=providers_dict.get(gs.provider_id, "") if gs.provider_id else "")
        ws.cell(row=row_num, column=11, value=gs.is_validated or "pending")
        ws.cell(row=row_num, column=12, value=gs.validation_errors or "")
    
    # Автоматическая ширина колонок
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Сохраняем в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info(
        "Экспорт АЗС завершен",
        extra={
            "total_gas_stations": total,
            "format": "xlsx"
        }
    )
    
    # Логируем действие пользователя
    if current_user:
        try:
            logging_service.log_user_action(
                db=db,
                user_id=current_user.id,
                username=current_user.username,
                action_type="export",
                action_description=f"Экспортированы АЗС",
                action_category="gas_station",
                entity_type="GasStation",
                entity_id=None,
                status="success",
                extra_data={
                    "total": total,
                    "is_validated": is_validated,
                    "provider_id": provider_id
                }
            )
        except Exception as e:
            logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
    
    # Определяем имя файла
    filename = f"gas_stations_export_{date.today().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/{gas_station_id}", response_model=GasStationResponse)
async def get_gas_station(
    gas_station_id: int,
    db: Session = Depends(get_db)
):
    """
    Получение АЗС по ID
    """
    gas_station_service = GasStationService(db)
    gas_station = gas_station_service.get_gas_station(gas_station_id)
    if not gas_station:
        raise HTTPException(status_code=404, detail="АЗС не найдена")
    return gas_station


@router.put("/{gas_station_id}", response_model=GasStationResponse)
async def update_gas_station(
    gas_station_id: int,
    gas_station_update: GasStationUpdate,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Обновление данных АЗС
    """
    gas_station_service = GasStationService(db)
    # original_name нельзя изменять - это поле только для чтения
    # Оно устанавливается только при создании записи из загружаемых файлов
    gas_station = gas_station_service.update_gas_station(
        gas_station_id=gas_station_id,
        original_name=None,  # Не изменяем original_name
        name=gas_station_update.name,  # Позволяем изменять наименование
        provider_id=gas_station_update.provider_id,
        azs_number=gas_station_update.azs_number,
        location=gas_station_update.location,
        region=gas_station_update.region,
        settlement=gas_station_update.settlement,
        latitude=gas_station_update.latitude,
        longitude=gas_station_update.longitude,
        is_validated=gas_station_update.is_validated
    )
    
    if not gas_station:
        raise HTTPException(status_code=404, detail="АЗС не найдена")
    
    logger.info("АЗС обновлена", extra={"gas_station_id": gas_station_id})
    
    # Логируем действие пользователя
    if current_user:
        try:
            logging_service.log_user_action(
                db=db,
                user_id=current_user.id,
                username=current_user.username,
                action_type="update",
                action_description=f"Обновлена АЗС: {gas_station.original_name}",
                action_category="gas_station",
                entity_type="GasStation",
                entity_id=gas_station_id,
                status="success"
            )
        except Exception as e:
            logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
    
    return gas_station


@router.delete("/{gas_station_id}")
async def delete_gas_station(
    gas_station_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Удаление АЗС по ID
    """
    gas_station_service = GasStationService(db)
    
    # Получаем информацию об АЗС перед удалением для логирования
    gas_station = gas_station_service.get_gas_station(gas_station_id)
    if not gas_station:
        raise HTTPException(status_code=404, detail="АЗС не найдена")
    
    gas_station_name = gas_station.name or gas_station.original_name
    
    # Удаляем АЗС
    success, error_message = gas_station_service.delete_gas_station(gas_station_id)
    
    if not success:
        if error_message:
            raise HTTPException(status_code=400, detail=error_message)
        else:
            raise HTTPException(status_code=404, detail="АЗС не найдена")
    
    logger.info("АЗС удалена", extra={"gas_station_id": gas_station_id})
    
    # Логируем действие пользователя
    if current_user:
        try:
            logging_service.log_user_action(
                db=db,
                user_id=current_user.id,
                username=current_user.username,
                action_type="delete",
                action_description=f"Удалена АЗС: {gas_station_name}",
                action_category="gas_station",
                entity_type="GasStation",
                entity_id=gas_station_id,
                status="success"
            )
        except Exception as e:
            logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
    
    return {"message": "АЗС успешно удалена"}


@router.delete("/clear")
async def clear_all_gas_stations(
    confirm: Optional[str] = Query(None, description="Подтверждение удаления всех АЗС"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_admin)
):
    """
    Очистка всех АЗС из базы данных
    """
    confirm_bool = confirm and confirm.lower() in ("true", "1", "yes")
    
    if not confirm_bool:
        raise HTTPException(
            status_code=400, 
            detail="Для очистки всех АЗС необходимо установить параметр confirm=true"
        )
    
    try:
        total_count = db.query(GasStation).count()
        db.query(GasStation).delete()
        db.commit()
        
        # Логируем действие пользователя
        if current_user:
            try:
                logging_service.log_user_action(
                    db=db,
                    user_id=current_user.id,
                    username=current_user.username,
                    action_type="clear",
                    action_description=f"Очищены все АЗС ({total_count} записей)",
                    action_category="gas_station",
                    entity_type="GasStation",
                    entity_id=None,
                    status="success",
                    extra_data={"deleted_count": total_count}
                )
            except Exception as e:
                logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
        
        logger.info(f"Очищены все АЗС", extra={"deleted_count": total_count})
        
        return {
            "message": f"Все АЗС успешно удалены",
            "deleted_count": total_count
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Ошибка при очистке АЗС", extra={"error": str(e)}, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ошибка при очистке АЗС: {str(e)}")


@router.post("/import")
@limiter.limit(settings.rate_limit_strict)
async def import_gas_stations(
    request: Request,
    file: UploadFile = File(...),
    provider_id: Optional[int] = Query(None, description="ID провайдера (если не указан, будет попытка определения из файла)"),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Импорт АЗС из Excel файла
    
    Ожидаемые колонки в Excel файле:
    - original_name (обязательно) - Исходное наименование АЗС
    - name (опционально) - Наименование АЗС (если не указано, будет равно original_name)
    - azs_number (опционально) - Номер АЗС
    - location (опционально) - Местоположение
    - region (опционально) - Регион
    - settlement (опционально) - Населенный пункт
    - latitude (опционально) - Широта
    - longitude (опционально) - Долгота
    - provider_id или provider_name (опционально) - ID или название провайдера
    
    Если provider_id не указан в параметрах, система попытается определить провайдера
    из колонки provider_id или provider_name в файле.
    """
    logger.info(
        f"Начало импорта АЗС из файла: {file.filename}",
        extra={
            "file_name": file.filename,
            "content_type": file.content_type,
            "provider_id": provider_id
        }
    )
    
    # Валидация типа файла
    try:
        validate_excel_file(file)
    except HTTPException as e:
        logger.error(
            "Ошибка валидации файла",
            extra={
                "file_name": file.filename,
                "error": str(e.detail),
                "status_code": e.status_code
            }
        )
        raise
    
    # Сохраняем файл во временную директорию
    tmp_file_path = None
    try:
        # Читаем содержимое файла
        content = await file.read()
        
        # Валидация размера файла
        from app.utils import validate_file_size
        validate_file_size(content, settings.max_file_size)
        
        # Создаем временный файл
        tmp_file_path = create_temp_file(content, file.filename)
        
        # Читаем Excel файл
        try:
            df = pd.read_excel(tmp_file_path, engine="openpyxl")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Ошибка чтения Excel файла: {str(e)}")
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Файл пуст или не содержит данных")
        
        # Нормализуем названия колонок (приводим к нижнему регистру, убираем пробелы)
        df.columns = df.columns.str.strip().str.lower()
        
        # Проверяем наличие обязательной колонки original_name
        if 'original_name' not in df.columns and 'исходное наименование' not in df.columns:
            # Пробуем найти колонку с похожим названием
            possible_names = ['наименование', 'название', 'name', 'азс']
            found_col = None
            for col in df.columns:
                for possible in possible_names:
                    if possible in col.lower():
                        found_col = col
                        break
                if found_col:
                    break
            
            if not found_col:
                raise HTTPException(
                    status_code=400,
                    detail="Не найдена обязательная колонка 'original_name' (Исходное наименование). "
                           "Пожалуйста, убедитесь, что файл содержит колонку с названием АЗС."
                )
            df.rename(columns={found_col: 'original_name'}, inplace=True)
        
        # Маппинг русских названий колонок на английские
        column_mapping = {
            'исходное наименование': 'original_name',
            'наименование': 'name',
            'номер азс': 'azs_number',
            'номер': 'azs_number',
            'местоположение': 'location',
            'регион': 'region',
            'населенный пункт': 'settlement',
            'широта': 'latitude',
            'долгота': 'longitude',
            'id провайдера': 'provider_id',
            'провайдер': 'provider_name',
            'название провайдера': 'provider_name'
        }
        
        for ru_name, en_name in column_mapping.items():
            if ru_name in df.columns and en_name not in df.columns:
                df.rename(columns={ru_name: en_name}, inplace=True)
        
        # Инициализируем сервис
        gas_station_service = GasStationService(db)
        
        # Статистика импорта
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        warnings = []
        
        # Обрабатываем каждую строку
        for idx, row in df.iterrows():
            try:
                # Получаем original_name (обязательное поле)
                original_name = row.get('original_name')
                if pd.isna(original_name) or not str(original_name).strip():
                    skipped_count += 1
                    errors.append(f"Строка {idx + 2}: отсутствует исходное наименование АЗС")
                    continue
                
                original_name = str(original_name).strip()
                
                # Получаем остальные поля
                name = row.get('name')
                if pd.isna(name):
                    name = None
                else:
                    name = str(name).strip() if name else None
                
                azs_number = row.get('azs_number')
                if pd.isna(azs_number):
                    azs_number = None
                else:
                    azs_number = str(azs_number).strip() if azs_number else None
                
                location = row.get('location')
                if pd.isna(location):
                    location = None
                else:
                    location = str(location).strip() if location else None
                
                region = row.get('region')
                if pd.isna(region):
                    region = None
                else:
                    region = str(region).strip() if region else None
                
                settlement = row.get('settlement')
                if pd.isna(settlement):
                    settlement = None
                else:
                    settlement = str(settlement).strip() if settlement else None
                
                latitude = row.get('latitude')
                if pd.isna(latitude):
                    latitude = None
                else:
                    try:
                        latitude = float(latitude)
                    except (ValueError, TypeError):
                        latitude = None
                
                longitude = row.get('longitude')
                if pd.isna(longitude):
                    longitude = None
                else:
                    try:
                        longitude = float(longitude)
                    except (ValueError, TypeError):
                        longitude = None
                
                # Определяем provider_id
                current_provider_id = provider_id
                
                # Если provider_id не указан в параметрах, пытаемся определить из файла
                if not current_provider_id:
                    file_provider_id = row.get('provider_id')
                    if not pd.isna(file_provider_id):
                        try:
                            current_provider_id = int(file_provider_id)
                        except (ValueError, TypeError):
                            pass
                    
                    # Если не нашли по ID, пытаемся найти по названию
                    if not current_provider_id:
                        provider_name = row.get('provider_name')
                        if not pd.isna(provider_name) and provider_name:
                            provider_name = str(provider_name).strip()
                            provider = db.query(Provider).filter(Provider.name == provider_name).first()
                            if provider:
                                current_provider_id = provider.id
                
                # Проверяем существование провайдера, если указан
                if current_provider_id:
                    provider = db.query(Provider).filter(Provider.id == current_provider_id).first()
                    if not provider:
                        warnings.append(f"Строка {idx + 2}: провайдер с ID {current_provider_id} не найден, АЗС будет создана без провайдера")
                        current_provider_id = None
                
                # Создаем или получаем АЗС
                gas_station, gas_station_warnings = gas_station_service.get_or_create_gas_station(
                    original_name=original_name,
                    azs_number=azs_number,
                    location=location,
                    region=region,
                    settlement=settlement,
                    latitude=latitude,
                    longitude=longitude,
                    provider_id=current_provider_id
                )
                
                # Обновляем name, если указано и отличается от original_name
                if name and name != gas_station.name:
                    gas_station.name = name
                    db.flush()
                
                # Проверяем, была ли АЗС создана или обновлена
                if gas_station.created_at and (datetime.now() - gas_station.created_at).total_seconds() < 5:
                    created_count += 1
                else:
                    updated_count += 1
                
                # Собираем предупреждения
                if gas_station_warnings:
                    warnings.extend([f"Строка {idx + 2}: {w}" for w in gas_station_warnings])
                
            except Exception as e:
                error_msg = f"Строка {idx + 2}: {str(e)}"
                errors.append(error_msg)
                logger.error(f"Ошибка при обработке строки {idx + 2}", extra={"error": str(e)}, exc_info=True)
                skipped_count += 1
        
        # Коммитим изменения
        db.commit()
        
        logger.info(
            "Импорт АЗС завершен",
            extra={
                "file_name": file.filename,
                "created_count": created_count,
                "updated_count": updated_count,
                "skipped_count": skipped_count,
                "errors_count": len(errors),
                "warnings_count": len(warnings)
            }
        )
        
        # Логируем действие пользователя
        if current_user:
            try:
                logging_service.log_user_action(
                    db=db,
                    user_id=current_user.id,
                    username=current_user.username,
                    action_type="import",
                    action_description=f"Импортированы АЗС из файла: {file.filename}",
                    action_category="gas_station",
                    entity_type="GasStation",
                    entity_id=None,
                    status="success",
                    extra_data={
                        "file_name": file.filename,
                        "created_count": created_count,
                        "updated_count": updated_count,
                        "skipped_count": skipped_count
                    }
                )
            except Exception as e:
                logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
        
        return {
            "message": "Импорт АЗС завершен",
            "created": created_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "errors": errors[:50],  # Ограничиваем количество ошибок в ответе
            "warnings": warnings[:50]  # Ограничиваем количество предупреждений в ответе
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Ошибка при импорте АЗС", extra={"error": str(e)}, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ошибка при импорте АЗС: {str(e)}")
    finally:
        # Удаляем временный файл
        if tmp_file_path:
            cleanup_temp_file(tmp_file_path)

