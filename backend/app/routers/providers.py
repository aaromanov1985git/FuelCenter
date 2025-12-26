"""
Роутер для работы с провайдерами
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.database import get_db
from app.logger import logger
from app.models import Provider, ProviderTemplate, User, Transaction, GasStation, FuelCard, Vehicle
from app.auth import require_auth_if_enabled
from app.services.logging_service import logging_service
from app.services.transaction_service import TransactionService
from app.schemas import (
    ProviderResponse, ProviderCreate, ProviderUpdate, ProviderListResponse,
    ProviderTemplateResponse, ProviderTemplateCreate, ProviderTemplateListResponse
)
from app.services.provider_service import ProviderService
from app.utils import serialize_template_json
from app.services.cache_service import CacheService
import hashlib
import json

router = APIRouter(prefix="/api/v1/providers", tags=["providers"])
cache = CacheService.get_instance()


@router.get("", response_model=ProviderListResponse)
async def get_providers(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    is_active: Optional[bool] = Query(None, description="Фильтр по активности"),
    organization_id: Optional[int] = Query(None, description="Фильтр по организации"),
    db: Session = Depends(get_db)
):
    """
    Получение списка провайдеров
    """
    # Создаем ключ кэша
    cache_key_data = {
        "skip": skip,
        "limit": limit,
        "is_active": is_active,
        "organization_id": organization_id
    }
    cache_key = hashlib.md5(json.dumps(cache_key_data, sort_keys=True).encode()).hexdigest()
    cache_key_full = f"providers:list:{cache_key}"
    
    # Пробуем получить из кэша (TTL 5 минут для справочников)
    cached_result = cache.get(cache_key_full, prefix="")
    if cached_result is not None:
        logger.debug("Cache hit для списка провайдеров", extra={"cache_key": cache_key})
        return ProviderListResponse(**cached_result)
    
    provider_service = ProviderService(db)
    providers, total = provider_service.get_providers(
        skip=skip,
        limit=limit,
        is_active=is_active,
        organization_id=organization_id
    )
    
    logger.debug("Список провайдеров загружен", extra={"total": total, "returned": len(providers)})
    
    result = ProviderListResponse(total=total, items=providers)
    
    # Кэшируем результат (5 минут)
    cache.set(
        cache_key_full,
        {"total": result.total, "items": [item.model_dump() for item in result.items]},
        ttl=300,
        prefix=""
    )
    logger.debug("Cache miss, сохранено в кэш", extra={"cache_key": cache_key})
    
    return result


@router.get("/{provider_id}", response_model=ProviderResponse)
async def get_provider(provider_id: int, db: Session = Depends(get_db)):
    """
    Получение провайдера по ID
    """
    provider_service = ProviderService(db)
    provider = provider_service.get_provider(provider_id)
    if not provider:
        raise HTTPException(status_code=404, detail="Провайдер не найден")
    return provider


@router.post("", response_model=ProviderResponse)
async def create_provider(
    provider: ProviderCreate,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Создание нового провайдера
    """
    provider_service = ProviderService(db)
    try:
        db_provider = provider_service.create_provider(
            name=provider.name,
            code=provider.code,
            organization_id=provider.organization_id,
            is_active=provider.is_active
        )
        
        # Логируем действие пользователя
        if current_user:
            try:
                logging_service.log_user_action(
                    db=db,
                    user_id=current_user.id,
                    username=current_user.username,
                    action_type="create",
                    action_description=f"Создан провайдер: {db_provider.name}",
                    action_category="provider",
                    entity_type="Provider",
                    entity_id=db_provider.id,
                    status="success"
                )
            except Exception as e:
                logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
        
        return db_provider
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/{provider_id}", response_model=ProviderResponse)
async def update_provider(
    provider_id: int,
    provider: ProviderUpdate,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Обновление провайдера
    """
    provider_service = ProviderService(db)
    try:
        db_provider = provider_service.update_provider(
            provider_id=provider_id,
            name=provider.name,
            code=provider.code,
            organization_id=provider.organization_id,
            is_active=provider.is_active
        )
        if not db_provider:
            raise HTTPException(status_code=404, detail="Провайдер не найден")
        
        # Логируем действие пользователя
        if current_user:
            try:
                logging_service.log_user_action(
                    db=db,
                    user_id=current_user.id,
                    username=current_user.username,
                    action_type="update",
                    action_description=f"Обновлен провайдер: {db_provider.name}",
                    action_category="provider",
                    entity_type="Provider",
                    entity_id=db_provider.id,
                    status="success"
                )
            except Exception as e:
                logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
        
        return db_provider
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/{provider_id}")
async def delete_provider(
    provider_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Удаление провайдера
    """
    provider_service = ProviderService(db)
    # Получаем информацию о провайдере перед удалением для логирования
    provider = provider_service.get_provider(provider_id)
    provider_name = provider.name if provider else f"ID {provider_id}"
    
    success = provider_service.delete_provider(provider_id)
    if not success:
        raise HTTPException(status_code=404, detail="Провайдер не найден")
    
    # Логируем действие пользователя
    if current_user:
        try:
            logging_service.log_user_action(
                db=db,
                user_id=current_user.id,
                username=current_user.username,
                action_type="delete",
                action_description=f"Удален провайдер: {provider_name}",
                action_category="provider",
                entity_type="Provider",
                entity_id=provider_id,
                status="success"
            )
        except Exception as e:
            logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
    
    return {"message": "Провайдер успешно удален"}


@router.get("/{provider_id}/templates", response_model=ProviderTemplateListResponse)
async def get_provider_templates(
    provider_id: int,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    is_active: Optional[bool] = Query(None, description="Фильтр по активности"),
    db: Session = Depends(get_db)
):
    """
    Получение списка шаблонов провайдера
    """
    provider_service = ProviderService(db)
    templates, total = provider_service.get_provider_templates(
        provider_id=provider_id,
        skip=skip,
        limit=limit,
        is_active=is_active
    )
    
    if total == 0 and not provider_service.get_provider(provider_id):
        raise HTTPException(status_code=404, detail="Провайдер не найден")
    
    logger.debug("Список шаблонов провайдера загружен", extra={"provider_id": provider_id, "total": total})
    
    return ProviderTemplateListResponse(total=total, items=templates)


@router.post("/{provider_id}/templates", response_model=ProviderTemplateResponse)
async def create_template(
    provider_id: int,
    template: ProviderTemplateCreate,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Создание нового шаблона провайдера
    """
    logger.info("Получен запрос на создание шаблона", extra={
        "provider_id": provider_id,
        "template_name": template.name,
        "connection_type": template.connection_type
    })
    
    provider = db.query(Provider).filter(Provider.id == provider_id).first()
    if not provider:
        raise HTTPException(status_code=404, detail="Провайдер не найден")
    
    # Убеждаемся, что provider_id в теле запроса совпадает с URL (если указан)
    if template.provider_id != provider_id:
        logger.warning("Несоответствие provider_id", extra={
            "url_provider_id": provider_id,
            "body_provider_id": template.provider_id
        })
        raise HTTPException(
            status_code=400,
            detail=f"provider_id в теле запроса ({template.provider_id}) не совпадает с ID в URL ({provider_id})"
        )
    
    # Конвертируем field_mapping в JSON строку
    field_mapping_json = serialize_template_json(template.field_mapping)
    
    # Конвертируем connection_settings в JSON строку, если указаны
    connection_settings_json = serialize_template_json(template.connection_settings)
    
    # Конвертируем fuel_type_mapping в JSON строку, если указано
    fuel_type_mapping_json = serialize_template_json(template.fuel_type_mapping) if template.fuel_type_mapping else None
    
    db_template = ProviderTemplate(
        provider_id=provider_id,
        name=template.name,
        description=template.description,
        connection_type=template.connection_type if template.connection_type else "file",
        connection_settings=connection_settings_json,
        field_mapping=field_mapping_json,
        header_row=template.header_row if template.header_row is not None else 0,
        data_start_row=template.data_start_row if template.data_start_row is not None else 1,
        source_table=template.source_table,
        source_query=template.source_query,
        fuel_type_mapping=fuel_type_mapping_json,
        is_active=template.is_active if template.is_active is not None else True,
        auto_load_enabled=template.auto_load_enabled if template.auto_load_enabled is not None else False,
        auto_load_schedule=template.auto_load_schedule,
        auto_load_date_from_offset=template.auto_load_date_from_offset if template.auto_load_date_from_offset is not None else -7,
        auto_load_date_to_offset=template.auto_load_date_to_offset if template.auto_load_date_to_offset is not None else -1
    )
    db.add(db_template)
    db.commit()
    db.refresh(db_template)
    
    logger.info("Шаблон создан", extra={"template_id": db_template.id, "provider_id": provider_id})
    
    # Логируем действие пользователя
    if current_user:
        try:
            logging_service.log_user_action(
                db=db,
                user_id=current_user.id,
                username=current_user.username,
                action_type="create",
                action_description=f"Создан шаблон: {db_template.name}",
                action_category="template",
                entity_type="ProviderTemplate",
                entity_id=db_template.id,
                status="success",
                extra_data={"provider_id": provider_id}
            )
        except Exception as e:
            logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
    
    # Перезагружаем расписания, если у нового шаблона включена автозагрузка
    if db_template.auto_load_enabled and db_template.auto_load_schedule:
        try:
            from app.services.scheduler_service import SchedulerService
            scheduler = SchedulerService.get_instance()
            scheduler.reload_schedules()
            logger.info("Расписания автоматической загрузки перезагружены после создания шаблона", extra={
                "template_id": db_template.id
            })
        except Exception as e:
            logger.warning("Не удалось перезагрузить расписания после создания шаблона", extra={
                "template_id": db_template.id,
                "error": str(e)
            })
    
    # Явно преобразуем объект SQLAlchemy в Pydantic модель для корректной сериализации
    try:
        template_response = ProviderTemplateResponse.model_validate(db_template)
        return template_response
    except Exception as e:
        logger.error(f"Ошибка при преобразовании шаблона в ответ: {e}", exc_info=True)
        # Если преобразование не удалось, возвращаем объект напрямую (FastAPI попытается сериализовать)
        return db_template


@router.get("/{provider_id}/export")
async def export_provider_data(
    provider_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(require_auth_if_enabled)
):
    """
    Экспорт всех данных по провайдеру в Excel файл
    
    Экспортирует:
    - Транзакции
    - АЗС
    - Топливные карты
    - Транспортные средства
    
    Все данные экспортируются в один Excel файл с несколькими листами.
    """
    # Проверяем существование провайдера
    provider = db.query(Provider).filter(Provider.id == provider_id).first()
    if not provider:
        raise HTTPException(status_code=404, detail="Провайдер не найден")
    
    logger.info(
        "Начало экспорта данных по провайдеру",
        extra={"provider_id": provider_id, "provider_name": provider.name}
    )
    
    # Создаем Excel файл
    wb = Workbook()
    
    # Удаляем дефолтный лист
    wb.remove(wb.active)
    
    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 1. Лист "Транзакции"
    ws_transactions = wb.create_sheet("Транзакции")
    transaction_service = TransactionService(db)
    transactions, total_transactions = transaction_service.get_transactions(
        skip=0,
        limit=100000,
        provider_id=provider_id,
        sort_by="transaction_date",
        sort_order="desc"
    )
    
    headers_transactions = [
        "ID", "Дата и время", "№ карты", "Закреплена за", "Номер АЗС",
        "Товар / услуга", "Тип операции", "Количество", "Валюта",
        "Цена", "Сумма", "Организация", "Источник"
    ]
    
    for col_num, header in enumerate(headers_transactions, 1):
        cell = ws_transactions.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_num, trans in enumerate(transactions, 2):
        ws_transactions.cell(row=row_num, column=1, value=trans.get("id"))
        ws_transactions.cell(row=row_num, column=2, value=trans.get("transaction_date").strftime("%d.%m.%Y %H:%M") if trans.get("transaction_date") else "")
        ws_transactions.cell(row=row_num, column=3, value=trans.get("card_number") or "")
        ws_transactions.cell(row=row_num, column=4, value=trans.get("vehicle_display_name") or trans.get("vehicle") or "")
        ws_transactions.cell(row=row_num, column=5, value=trans.get("azs_number") or "")
        ws_transactions.cell(row=row_num, column=6, value=trans.get("product") or "")
        ws_transactions.cell(row=row_num, column=7, value=trans.get("operation_type") or "Покупка")
        ws_transactions.cell(row=row_num, column=8, value=float(trans.get("quantity", 0)) if trans.get("quantity") else "")
        ws_transactions.cell(row=row_num, column=9, value=trans.get("currency") or "RUB")
        ws_transactions.cell(row=row_num, column=10, value=float(trans.get("price", 0)) if trans.get("price") else "")
        ws_transactions.cell(row=row_num, column=11, value=float(trans.get("amount", 0)) if trans.get("amount") else "")
        ws_transactions.cell(row=row_num, column=12, value=trans.get("organization") or "")
        ws_transactions.cell(row=row_num, column=13, value=trans.get("source_file") or "")
    
    # Автоматическая ширина колонок для транзакций
    for col in ws_transactions.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_transactions.column_dimensions[col_letter].width = adjusted_width
    
    # 2. Лист "АЗС"
    ws_gas_stations = wb.create_sheet("АЗС")
    gas_stations = db.query(GasStation).filter(GasStation.provider_id == provider_id).all()
    
    headers_gas_stations = [
        "ID", "Исходное наименование", "Наименование", "Номер АЗС",
        "Местоположение", "Регион", "Населенный пункт",
        "Широта", "Долгота", "Статус валидации", "Ошибки валидации"
    ]
    
    for col_num, header in enumerate(headers_gas_stations, 1):
        cell = ws_gas_stations.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_num, gs in enumerate(gas_stations, 2):
        ws_gas_stations.cell(row=row_num, column=1, value=gs.id)
        ws_gas_stations.cell(row=row_num, column=2, value=gs.original_name or "")
        ws_gas_stations.cell(row=row_num, column=3, value=gs.name or "")
        ws_gas_stations.cell(row=row_num, column=4, value=gs.azs_number or "")
        ws_gas_stations.cell(row=row_num, column=5, value=gs.location or "")
        ws_gas_stations.cell(row=row_num, column=6, value=gs.region or "")
        ws_gas_stations.cell(row=row_num, column=7, value=gs.settlement or "")
        ws_gas_stations.cell(row=row_num, column=8, value=float(gs.latitude) if gs.latitude else "")
        ws_gas_stations.cell(row=row_num, column=9, value=float(gs.longitude) if gs.longitude else "")
        ws_gas_stations.cell(row=row_num, column=10, value=gs.is_validated or "pending")
        ws_gas_stations.cell(row=row_num, column=11, value=gs.validation_errors or "")
    
    # Автоматическая ширина колонок для АЗС
    for col in ws_gas_stations.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_gas_stations.column_dimensions[col_letter].width = adjusted_width
    
    # 3. Лист "Топливные карты"
    ws_fuel_cards = wb.create_sheet("Топливные карты")
    fuel_cards = db.query(FuelCard).filter(FuelCard.provider_id == provider_id).all()
    
    headers_fuel_cards = [
        "ID", "Номер карты", "Владелец (исходное)", "Владелец (нормализованное)",
        "Закреплена за ТС", "Дата начала закрепления", "Дата окончания закрепления",
        "Активное закрепление", "Заблокирована"
    ]
    
    for col_num, header in enumerate(headers_fuel_cards, 1):
        cell = ws_fuel_cards.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_num, card in enumerate(fuel_cards, 2):
        ws_fuel_cards.cell(row=row_num, column=1, value=card.id)
        ws_fuel_cards.cell(row=row_num, column=2, value=card.card_number or "")
        ws_fuel_cards.cell(row=row_num, column=3, value=card.original_owner_name or "")
        ws_fuel_cards.cell(row=row_num, column=4, value=card.normalized_owner or "")
        # Получаем название ТС, если есть
        vehicle_name = ""
        if card.vehicle_id:
            vehicle = db.query(Vehicle).filter(Vehicle.id == card.vehicle_id).first()
            if vehicle:
                vehicle_name = vehicle.original_name
        ws_fuel_cards.cell(row=row_num, column=5, value=vehicle_name)
        ws_fuel_cards.cell(row=row_num, column=6, value=card.assignment_start_date.strftime("%d.%m.%Y") if card.assignment_start_date else "")
        ws_fuel_cards.cell(row=row_num, column=7, value=card.assignment_end_date.strftime("%d.%m.%Y") if card.assignment_end_date else "")
        ws_fuel_cards.cell(row=row_num, column=8, value="Да" if card.is_active_assignment else "Нет")
        ws_fuel_cards.cell(row=row_num, column=9, value="Да" if card.is_blocked else "Нет")
    
    # Автоматическая ширина колонок для карт
    for col in ws_fuel_cards.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_fuel_cards.column_dimensions[col_letter].width = adjusted_width
    
    # 4. Лист "Транспортные средства"
    # Получаем уникальные ТС из транзакций провайдера
    vehicle_ids = db.query(Transaction.vehicle_id).filter(
        Transaction.provider_id == provider_id,
        Transaction.vehicle_id.isnot(None)
    ).distinct().all()
    vehicle_ids = [v[0] for v in vehicle_ids]
    
    ws_vehicles = wb.create_sheet("Транспортные средства")
    vehicles = db.query(Vehicle).filter(Vehicle.id.in_(vehicle_ids)).all() if vehicle_ids else []
    
    headers_vehicles = [
        "ID", "Исходное наименование", "Гаражный номер", "Государственный номер",
        "Статус валидации", "Ошибки валидации"
    ]
    
    for col_num, header in enumerate(headers_vehicles, 1):
        cell = ws_vehicles.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_num, vehicle in enumerate(vehicles, 2):
        ws_vehicles.cell(row=row_num, column=1, value=vehicle.id)
        ws_vehicles.cell(row=row_num, column=2, value=vehicle.original_name or "")
        ws_vehicles.cell(row=row_num, column=3, value=vehicle.garage_number or "")
        ws_vehicles.cell(row=row_num, column=4, value=vehicle.license_plate or "")
        ws_vehicles.cell(row=row_num, column=5, value=vehicle.is_validated or "pending")
        ws_vehicles.cell(row=row_num, column=6, value=vehicle.validation_errors or "")
    
    # Автоматическая ширина колонок для ТС
    for col in ws_vehicles.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_vehicles.column_dimensions[col_letter].width = adjusted_width
    
    # Сохраняем в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info(
        "Экспорт данных по провайдеру завершен",
        extra={
            "provider_id": provider_id,
            "provider_name": provider.name,
            "transactions_count": total_transactions,
            "gas_stations_count": len(gas_stations),
            "fuel_cards_count": len(fuel_cards),
            "vehicles_count": len(vehicles)
        }
    )
    
    # Логируем действие пользователя
    if current_user:
        try:
            logging_service.log_user_action(
                db=db,
                user_id=current_user.id,
                username=current_user.username,
                action_type="export",
                action_description=f"Экспортированы данные провайдера: {provider.name}",
                action_category="provider",
                entity_type="Provider",
                entity_id=provider_id,
                status="success",
                extra_data={
                    "transactions_count": total_transactions,
                    "gas_stations_count": len(gas_stations),
                    "fuel_cards_count": len(fuel_cards),
                    "vehicles_count": len(vehicles)
                }
            )
        except Exception as e:
            logger.error(f"Ошибка при логировании действия пользователя: {e}", exc_info=True)
    
    # Определяем имя файла
    provider_name_safe = "".join(c for c in provider.name if c.isalnum() or c in (' ', '-', '_')).strip()
    filename = f"provider_{provider_name_safe}_{date.today().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
